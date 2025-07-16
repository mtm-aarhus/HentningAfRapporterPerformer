import json
from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from office365.runtime.auth.user_credential import UserCredential
from office365.sharepoint.client_context import ClientContext
import os
from OpenOrchestrator.database.queues import QueueElement
from datetime import datetime
import calendar
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
import time
from urllib.parse import urlparse, parse_qs, unquote
import string
import random 
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta  # kræver `python-dateutil`
import gc
import subprocess
import win32com.client as win32
from OpenOrchestrator.database.queues import QueueElement
import subprocess
import random 
import string
import pandas as pd
from openpyxl import load_workbook
import pandas as pd



orchestrator_connection = OrchestratorConnection("PythonOpusBookMark", os.getenv('OpenOrchestratorSQL'),os.getenv('OpenOrchestratorKey'), None)
conversion_in_progress = set()


def convert_xls_to_xlsx(path: str) -> None:
    import shutil
    absolute_path = os.path.abspath(path)
    if absolute_path in conversion_in_progress:
        print(f"Conversion already in progress for {absolute_path}. Skipping.")
        return
    
    conversion_in_progress.add(absolute_path)
    try:
        print(f'Absolute path {absolute_path} found')
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(absolute_path)
        wb.Sheets(1).Name = "YKMD_STD"

        new_path = os.path.splitext(absolute_path)[0] + ".xlsx"
        wb.SaveAs(new_path, FileFormat=51)
        wb.Close()
        excel.Application.Quit()
        del wb
        del excel
    except AttributeError as e:
        if "CLSIDToClassMap" in str(e):
            print("Corrupt gen_py detected, clearing cache...")
            shutil.rmtree(win32.gencache.GetGeneratePath(), ignore_errors=True)
            # prøv igen
            return convert_xls_to_xlsx(path)
        raise e
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        raise e
    finally:
        conversion_in_progress.remove(absolute_path)

queue_element = orchestrator_connection.get_next_queue_element('HentningAfRapporterQueue')
specific_content = json.loads(queue_element.data)

# Assign variables from SpecificContent
Navn = specific_content.get("Navn", None)
Sti = specific_content.get("Sti", None)
QueueName = specific_content.get("QueueName", None)
SharePointURL = specific_content.get("SharePointMappeLink", None)


# Opus bruger
OpusLogin = orchestrator_connection.get_credential("OpusBruger")
OpusUser = OpusLogin.username
OpusPassword = OpusLogin.password 

# Robotpassword
RobotCredential = orchestrator_connection.get_credential("Robot365User") 
RobotUsername = RobotCredential.username
RobotPassword = RobotCredential.password

print('Started process "Hentning af rapporter"')

#slet filen, hvis den ikke er slettet fra sidst
IllegalFileName = ["YKMD_STD.xls", "run1file.xls", "run2file.xls", "run3file.xls", "run4file.xls", "YKMD_STD.xlsx", "run1file.xlsx", "run2file.xlsx", "run3file.xlsx", "run4file.xlsx", "samlet.xlsx", "samlet_pænt.xlsx", "PSA011, Mangler Timeregistrering.xlsx "]
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
def delete_files():
    for filename in IllegalFileName:
        if os.path.exists(downloads_folder + '\\' + filename):
            os.remove(downloads_folder + '\\' + filename)
            print(f'{filename} removed')
delete_files()
# Start chrome
chrome_options = Options()
chrome_options.add_argument('--remote-debugging-pipe')
chrome_options.add_argument("--headless=new")  # More stable headless mode
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--safebrowsing-disable-download-protection")

chrome_options.add_experimental_option("prefs", {
    "download.default_directory": downloads_folder,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True  # <- stadig "True", da vi har slået den aktive beskyttelse fra med flaget
})

chrome_service = Service()
try:
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
except Exception as e:
    print(e)

#Log ind på opus
print("Navigating to Opus login page")
driver.get(orchestrator_connection.get_constant("OpusAdgangUrl").value)
WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "logonuidfield")))

driver.find_element(By.ID, "logonuidfield").send_keys(OpusUser)
driver.find_element(By.ID, "logonpassfield").send_keys(OpusPassword)
driver.find_element(By.ID, "buttonLogon").click()

print("Logged in to Opus portal successfully")

#Videre til rapportside
if QueueName == "option1":
    runs = [
        {
            "name": "run1",
            "start": (datetime.now().replace(day=1) - relativedelta(months=3)).strftime("%d.%m.%Y"),
            "end": (datetime.now().replace(day=1) - relativedelta(months=2) - timedelta(days=1)).strftime("%d.%m.%Y"),
            "Filename": "run1file"

        },
        {
            "name": "run2",
            "start": (datetime(datetime.now().year, datetime.now().month, 1) - relativedelta(months=2)).strftime("%d.%m.%Y"),
            "end": (datetime(datetime.now().year, datetime.now().month, 1) - relativedelta(months=1) - timedelta(days=1)).strftime("%d.%m.%Y"),
            "Filename": "run2file"
        },
        {
            "name": "run3",
            "start": (datetime(datetime.now().year, datetime.now().month, 1) - relativedelta(months=1)).strftime("%d.%m.%Y"),
            "end": (datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)).strftime("%d.%m.%Y"),
            "Filename": "run3file"
        }
    ]
elif QueueName == "option2":
    runs = [
        {
            "name": "run1",
            "start": (datetime.now().replace(day=1) - relativedelta(months=3)).strftime("%d.%m.%Y"),
            "end": (datetime.now().replace(day=1) - relativedelta(months=2) - timedelta(days=1)).strftime("%d.%m.%Y"),
            "Filename": "run1file"

        },
        {
            "name": "run2",
            "start": (datetime(datetime.now().year, datetime.now().month, 1) - relativedelta(months=2)).strftime("%d.%m.%Y"),
            "end": (datetime(datetime.now().year, datetime.now().month, 1) - relativedelta(months=1) - timedelta(days=1)).strftime("%d.%m.%Y"),
            "Filename": "run2file"
        },
        {
            "name": "run3",
            "start": (datetime(datetime.now().year, datetime.now().month, 1) - relativedelta(months=1)).strftime("%d.%m.%Y"),
            "end": (datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)).strftime("%d.%m.%Y"),
            "Filename": "run3file"
        }

    ]
else:
    runs = None


for run in runs:
    print(f'At run {run["Filename"]}')
    WebDriverWait(driver,10)

    for attempt in range(3):
        try:
            print(f'Forsøg {attempt+1}')
            driver.get(Sti)  # Naviger hver gang
            
            wait = WebDriverWait(driver, 30)
            
            wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
            wait.until(lambda d: d.current_url.startswith(Sti))
            
            print("HTML er klar og URL er korrekt")
            break
        except Exception as e:
            print(f'Fejl på forsøg {attempt+1}: {e}')
    else:
        raise Exception(f"Lykkedes ikke at nå til {Sti} efter 3 forsøg")


    wait = WebDriverWait(driver, 200)
    FileName = run['Filename']
    print(run['name'])
    try:
        # Vent på og skift til iframen
        wait.until(EC.frame_to_be_available_and_switch_to_it(
            (By.CSS_SELECTOR, "iframe[id^='iframe_Roundtrip']")
        ))
        # Vent på og klik på knappen
        element = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//a[@class='urBtnStd' and normalize-space(text())='Variabelskærm']"
        )))
        element.click()
        print('Variabelskærmknap klikket')


        # Vent på og find inputfeltet ved siden af 'Arbejdsdato'
        arbejdsdato_input = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//span[contains(text(), 'Arbejdsdato')]/ancestor::td/following-sibling::td//input[@type='text']"
            ))
        )

        initial_file_count = len(os.listdir(downloads_folder))
        print('Initial file count', initial_file_count)

        # Brug feltet
        arbejdsdato_input.clear()
        arbejdsdato_input.send_keys(f"{run["start"]} - {run["end"]}")
        ok_knap = wait.until(EC.element_to_be_clickable((By.ID, "DLG_VARIABLE_dlgBase_BTNOK")))
        ok_knap.click()

        print('looking for export button')
        WebDriverWait(driver, timeout = 60*15).until(EC.presence_of_element_located((By.ID, "BUTTON_EXPORT_btn1_acButton")))
        driver.find_element(By.ID, "BUTTON_EXPORT_btn1_acButton").click()
        print('clicked export button')
        print("Waiting for file download to complete")

        start_time = time.time()
        while True:
            files = os.listdir(downloads_folder)
            if len(files) > initial_file_count:
                latest_file = max(
                    [os.path.join(downloads_folder, f) for f in files], key=os.path.getctime
                )
                if latest_file.endswith(".xls"):
                    print('Found xls file')
                    new_file_path = os.path.join(downloads_folder, f"{FileName}.xls")
                    os.rename(latest_file, new_file_path)
                    print(f"File downloaded and renamed to {new_file_path}")
                    xlsx_file_path_check = True
                    break
                
            if time.time() - start_time > 3600:
                print("Mail sent due to timeout")
                raise TimeoutError("File download did not complete within 60 minutes.")
            time.sleep(1)
        for handle in driver.window_handles:
            if handle != driver.current_window_handle:
                driver.switch_to.window(handle)
                driver.close()
                print("Closed extra window")

        driver.switch_to.window(driver.window_handles[0])

        if xlsx_file_path_check:
            xlsx_file_path = os.path.join(downloads_folder, FileName + ".xlsx")
            try:
                print(f'Converting {new_file_path}')
                convert_xls_to_xlsx( new_file_path)
            except TimeoutError:
                orchestrator_connection.log_error(f'Conversion of {new_file_path} timed out')
    
            except Exception as e:
                gc.collect()
                subprocess.call("taskkill /im excel.exe /f >nul 2>&1", shell=True)
                time.sleep(2)
                if os.path.exists(xlsx_file_path):
                    os.remove(xlsx_file_path)
                orchestrator_connection.log_error(f'An error happened {str(e)}')
                raise e

    except Exception as e:
        orchestrator_connection.log_error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")
        driver.quit()
        raise e

    except Exception as e:
        print(f'Der skete en fejl: {e}')
driver.quit()
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
file_paths = [os.path.join(downloads_folder, run["Filename"]) + '.xlsx' for run in runs]

def find_table_start(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Operationsnr. og tekst":
                return cell.row - 1  # pandas er 0-indekseret
    raise ValueError(f"Tabelstart ikke fundet i {file}")

dfs = []

for i, file in enumerate(file_paths):
    start_row = find_table_start(file)
    df = pd.read_excel(file, skiprows=start_row)

    # Fjern header fra efterfølgende filer
    if i > 0:
        df = df.iloc[1:]

    # Behold kolonner, også dem med tom titel
    df.columns = [col if col else "" for col in df.columns]

    # Rens evt. helt tomme rækker
    df = df.dropna(how='all')

    dfs.append(df)

# Saml alt i én dataframe
result = pd.concat(dfs, ignore_index=True)
kolonnenavne = [
    "Organisatorisk enhed",
    "Profitcenter",
    "Profitctr. nr.",
    "Medarbejder",
    "Medarbejdernr.",
    "Arbejdsdato",
    "Operationsnr. og tekst",
    "Operationsnr.",
    "Aktivitetsart",
    "Akt. art nr.",
    "Arbejdstid",
    "Planlagt arbejdstid",
    "Fravær",
    "Overarbejde",
    "Tilstedevær",
    "Faktisk Arbejde",
    "Difference Arbejdstid",
    "% arb. af arbejdstid"
]

# efter du har samlet og evt. renset dine dataframes:
result.columns = kolonnenavne

# Skriv til pænt Excel
excel_file_path = "PSA011, Mangler Timeregistrering.xlsx"
with pd.ExcelWriter(excel_file_path, engine='xlsxwriter', datetime_format='dd-mm-yyyy') as writer:
    result.to_excel(writer, sheet_name='YKMD_STD', index=False)

    workbook  = writer.book
    worksheet = writer.sheets['Data']

    # Lav Excel-tabel over hele området
    (max_row, max_col) = result.shape
    col_range = chr(65 + max_col - 1)  # A … Z
    table_range = f"A1:{col_range}{max_row + 1}"

    worksheet.add_table(table_range, {
        'name': 'SamletTabel',
        'columns': [{'header': col if col else ""} for col in result.columns]
    })

    # Tilpas kolonnebredder
    for i, column in enumerate(result.columns):
        # find max længde i kolonnen (inkl. header)
        max_len = max(
            result[column].astype(str).map(len).max(),
            len(str(column))
        )
        worksheet.set_column(i, i, max_len + 2)

parsed_url = urlparse(SharePointURL)
base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
# **Automatically Detect if it's a Teams or Sites URL**
if "/Teams/" in SharePointURL:
    teamsite = SharePointURL.split('Teams/')[1].split('/')[0]
    base_url = f"{base_url}/Teams/{teamsite}"
elif "/Sites/" in SharePointURL:
    sitename = SharePointURL.split('Sites/')[1].split('/')[0]
    base_url = f"{base_url}/Sites/{sitename}"
else:
    print("WARNING: Could not determine if this is a Teams or Sites URL. Using default base_url.")
credentials = UserCredential(RobotUsername,RobotPassword)
ctx = ClientContext(base_url).with_credentials(credentials)



# Extract path correctly
query_params = parse_qs(parsed_url.query)
id_param = query_params.get("id", [None])[0]

if id_param:
    # If it's a sharing link with an ID, extract the correct path
    decoded_path = unquote(id_param).rstrip('/')
else:
    # Normal URL or sharing link without ID
    if "/r/" in SharePointURL:
        decoded_path = SharePointURL.split('/r/', 1)[1].split('?', 1)[0]
    else:
        decoded_path = parsed_url.path.lstrip('/')

# **Replace %20 with spaces to match SharePoint folder structure**
decoded_path = decoded_path.replace("%20", " ")

# Ensure the correct format
if not decoded_path.startswith("/"):
    decoded_path = "/" + decoded_path

folder_relative_url = decoded_path
target_folder = ctx.web.get_folder_by_server_relative_path(folder_relative_url)
ctx.load(target_folder)
ctx.execute_query()

# Upload file
file_name = os.path.basename("PSA011, Mangler Timeregistrering.xlsx")

with open("PSA011, Mangler Timeregistrering.xlsx", "rb") as local_file:
    target_folder.upload_file(file_name, local_file.read()).execute_query()
    
delete_files()